# coding: utf-8
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from openpyxl.styles import PatternFill, NamedStyle, Font, Side, Border


class Excel:
    def read_excel(self, filename):
        wb = load_workbook(filename)
        sheet = wb[wb.sheetnames[0]]
        return sheet

    def get_orders(self, sheet):
        orders = []
        for row in sheet.iter_rows(min_row=2):
            order = dict()
            order["item"] = str(row[0].value)
            order["friend_phone"] = str(row[1].value)
            order["name"] = row[2].value
            order["tel"] = str(row[3].value)
            order["address"] = row[4].value
            order["goods"] = []
            for i in range(5, len(row) - 1, 4):
                if row[i].value and row[i+1].value and row[i+2].value:
                    order["goods"].append(
                        {
                            "abiid": row[i].value,
                            "good_name": row[i+1].value,
                            "num": row[i+2].value,
                            "price": row[i+3].value
                        }
                    )
            orders.append(order)
        return orders

    def myorders(self, filename):
        try:
            sheet = self.read_excel(filename)
        except Exception as e:
            return {
                "code": 1,
                "msg": "读取excel失败,请检查文件, {}\n".format(e)
            }
        try:
            orders = self.get_orders(sheet)
        except Exception as e:
            return {
                "code": 1,
                "msg": "读取订单失败,excel格式错误, {}\n".format(e)
            }
        return {
            "msg": "读取订单成功\n",
            "code": 0,
            "data": orders
        }

    def save_orders(self, orders):
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        errors = {
            1: "地址错误",
            2: "电话号码错误",
            3: "商品id错误",
            4: "下单失败, 商品可能不存在或缺货",
            5: "委托人号码错误",
            6: "下单失败, 服务器错误",
            7: "网络错误",
            8: "商品价格错误"
        }
        wb = Workbook()
        wb.add_named_style(highlight)
        ws = wb.active
        ws.append(["error", "item", "委托人", "姓名", "电话", "地址", "商品id", "商品名称", "数量", "商品单价"])
        row = 2
        for order in orders:
            data = [
                order.get("item"),
                order.get("friend_phone"),
                order.get("name"),
                order.get("tel"),
                order.get("address")
            ]
            goods = order.get("goods")
            for good in goods:
                data += [good.get("abiid"), good.get("good_name"), good.get("num"), good.get("price")]
            error = order.get("error")
            if error:
                data.insert(0, errors.get(error))
            else:
                data.insert(0, "下单成功")
            ws.append(data)
            high_fill = PatternFill(fill_type='solid', fgColor="28ff86")
            if error == 1:
                ws["F" + str(row)].fill = high_fill
            elif error == 2:
                ws["E" + str(row)].fill = high_fill
            elif error == 3:
                ws["G" + str(row)].fill = high_fill
            elif error == 4:
                ws["A" + str(row)].fill = high_fill
            elif error == 5:
                ws["C" + str(row)].fill = high_fill
            elif error == 6:
                ws["A" + str(row)].fill = high_fill
            elif error == 7:
                ws["A" + str(row)].fill = high_fill
            elif error == 8:
                ws["J" + str(row)].fill = high_fill
            row += 1
        if os.path.exists("out.xlsx"):
            try:
                os.remove("out.xlsx")
            except Exception as e:
                return False
        wb.save("out.xlsx")
        return True

    def save_auther_orders(self, orders):
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        wb = Workbook()
        wb.add_named_style(highlight)
        ws = wb.active
        ws.append(["姓名", "电话", "地址", "商品id", "商品名称", "数量", "商品单价"])
        row = 2
        for order in orders:
            data = [
                order.get("name"),
                order.get("tel"),
                order.get("address"),
            ]
            goods = order.get("goods")
            for good in goods:
                data += [good.get("abiid"), good.get("good_name"), good.get("num"), good.get("price")]
            ws.append(data)
            high_fill = PatternFill(fill_type='solid', fgColor="28ff86")
            errors = order.get("errors")
            if errors.get("tel_error"):
                ws["B" + str(row)].fill = high_fill

            if errors.get("address_error"):
                ws["C" + str(row)].fill = high_fill

            if errors.get("good_error"):
                ws["D" + str(row)].fill = high_fill

            if errors.get("price_error"):
                ws["G" + str(row)].fill = high_fill

            row += 1
        if os.path.exists("out.xlsx"):
            try:
                os.remove("out.xlsx")
            except Exception as e:
                return False
        wb.save("out.xlsx")
        return True

    def get_auther_orders(self, sheet):
        orders = []
        for row in sheet.iter_rows(min_row=2):
            order = dict()
            order["name"] = str(row[0].value)
            order["tel"] = str(row[1].value)
            order["address"] = row[2].value
            order["goods"] = []
            for i in range(3, len(row) - 1, 4):
                if row[i].value and row[i+1].value and row[i+2].value:
                    order["goods"].append(
                        {
                            "abiid": row[i].value,
                            "good_name": row[i+1].value,
                            "num": row[i+2].value,
                            "price": row[i+3].value
                        }
                    )
            orders.append(order)
        return [order for order in orders if order.get("goods")]


